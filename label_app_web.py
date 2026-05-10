"""
Rotul — Versión Web (Streamlit)
Ejecutar con: streamlit run label_app_web.py
"""

import streamlit as st
import streamlit.components.v1 as _st_components
import base64
try:
    from streamlit_paste_button import paste_image_button as _paste_btn
    PASTE_BTN_OK = True
except ImportError:
    PASTE_BTN_OK = False
import json
import os
import sys
import uuid
import calendar
import datetime
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont

# ── ReportLab (PDF) ───────────────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4, portrait
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.utils import ImageReader
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False
    mm = 2.8346

# ── Rutas de datos ────────────────────────────────────────────────────────────
SCRIPT_DIR            = os.path.dirname(os.path.abspath(__file__))
IMAGE_DIR             = os.path.join(SCRIPT_DIR, "label_images")
DATA_FILE             = os.path.join(SCRIPT_DIR, "label_data.json")
FIELD_CONFIG_PATH     = os.path.join(SCRIPT_DIR, "field_config.json")
MATERIALS_CONFIG_PATH = os.path.join(SCRIPT_DIR, "materials_config.json")
HISTORY_PATH          = os.path.join(SCRIPT_DIR, "import_history.json")
GLOBAL_SUPPLIER_PATH  = os.path.join(IMAGE_DIR,  "global_supplier.png")
os.makedirs(IMAGE_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE PÁGINA
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Rotul",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; }
    .stButton>button { width: 100%; }
    div[data-testid="stSidebarContent"] { padding-top: 1rem; }
    .label-card {
        border: 1px solid #ddd; border-radius: 6px;
        padding: 8px; margin-bottom: 6px;
        background: #fafafa; cursor: pointer;
    }
    .label-card:hover { background: #f0f4ff; border-color: #336699; }
    .label-card.selected { background: #e8f0fe; border-color: #336699; border-width: 2px; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PERSISTENCIA EN SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
def _load_json(path, default):
    try:
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return default

def _save_json(path, data):
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        st.error(f"Error guardando {path}: {e}")

def _get_items():
    """Siempre devuelve una lista limpia de dicts desde session_state."""
    raw = st.session_state.get('_label_items', [])
    if callable(raw):
        raw = []
    if isinstance(raw, dict):
        raw = raw.get('items', [])
    if not isinstance(raw, list):
        raw = []
    return [i for i in raw if isinstance(i, dict)]

def _set_items(lst):
    """Guarda la lista en session_state de forma segura."""
    st.session_state['_label_items'] = [i for i in lst if isinstance(i, dict)]

def _init_state():
    if '_label_items' not in st.session_state or callable(st.session_state.get('_label_items')):
        raw = _load_json(DATA_FILE, {'items': []})
        if isinstance(raw, dict):
            items = raw.get('items', [])
        elif isinstance(raw, list):
            items = raw
        else:
            items = []
        _set_items(items)

    if 'field_config' not in st.session_state:
        default_fc = {
            'product':    {'label': 'PRODUCTO',   'visible': True},
            'oc':         {'label': 'OC',          'visible': True},
            'fab_date':   {'label': 'F.ELAB.',     'visible': True},
            'exp_date':   {'label': 'F.EXP.',      'visible': True},
            'batch':      {'label': 'Lote',        'visible': True},
            'quantity':   {'label': 'Cantidad',    'visible': True},
            'deliver_to': {'label': 'Entregar a',  'visible': True},
            'description':{'label': 'Descripción', 'visible': False},
            'etq_x_caja': {'label': 'ETQ x Caja', 'visible': False},
        }
        loaded = _load_json(FIELD_CONFIG_PATH, default_fc)
        # Migrar formato antiguo y agregar defaults faltantes
        for k, v in default_fc.items():
            if k not in loaded:
                loaded[k] = v
            elif not isinstance(loaded[k], dict):
                loaded[k] = v
            else:
                # Asegurar que tenga ambas keys
                if 'visible' not in loaded[k]: loaded[k]['visible'] = v['visible']
                if 'label'   not in loaded[k]: loaded[k]['label']   = v['label']
        # Forzar campos esenciales siempre visibles
        for essential in ['product','oc','fab_date','exp_date','batch','quantity','deliver_to']:
            if essential in loaded and not loaded[essential].get('visible', True):
                pass  # Respetar la config del usuario
        st.session_state.field_config = loaded

    if 'materials' not in st.session_state:
        st.session_state.materials = _load_json(MATERIALS_CONFIG_PATH, {})

    if 'global_sup_bytes' not in st.session_state:
        if os.path.exists(GLOBAL_SUPPLIER_PATH):
            with open(GLOBAL_SUPPLIER_PATH, 'rb') as f:
                st.session_state.global_sup_bytes = f.read()
        else:
            st.session_state.global_sup_bytes = None

    if 'selected_idx' not in st.session_state:
        st.session_state.selected_idx = None

    if 'search_query' not in st.session_state:
        st.session_state.search_query = ""

    if 'preview_page' not in st.session_state:
        st.session_state.preview_page = 0

_init_state()

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS — FUENTES Y DIBUJO
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_resource
def _get_font(size, bold=False):
    candidates_bold   = ["DejaVuSans-Bold.ttf","arialbd.ttf","Arial Bold.ttf","DejaVuSans.ttf","arial.ttf"]
    candidates_normal = ["DejaVuSans.ttf","arial.ttf","Arial.ttf","LiberationSans-Regular.ttf"]
    font_dirs = [
        "/usr/share/fonts/truetype/dejavu",
        "/usr/share/fonts/truetype/liberation",
        "/usr/share/fonts/truetype",
        "/usr/share/fonts",
        os.path.join(os.environ.get("WINDIR","C:/Windows"), "Fonts"),
    ]
    candidates = candidates_bold if bold else candidates_normal
    for c in candidates:
        try:
            return ImageFont.truetype(c, size)
        except Exception:
            pass
        for d in font_dirs:
            p = os.path.join(d, c)
            if os.path.exists(p):
                try:
                    return ImageFont.truetype(p, size)
                except Exception:
                    pass
    try:
        return ImageFont.load_default(size=size)
    except Exception:
        return ImageFont.load_default()

def _line_h(draw_obj, text, font):
    bb = font.getbbox(text or "A")
    return bb[3] - bb[1]

def _wrap_px(draw_obj, text, font, max_px):
    """Wrap text by pixel width, breaking long words if needed."""
    max_px = max(1, max_px)
    def split_word(w):
        chunks, chunk = [], ""
        for ch in w:
            if draw_obj.textlength(chunk+ch, font=font) <= max_px:
                chunk += ch
            else:
                if chunk: chunks.append(chunk)
                chunk = ch
        if chunk: chunks.append(chunk)
        return chunks or [w]

    words, lines, current = text.split(), [], ""
    for word in words:
        if draw_obj.textlength(word, font=font) > max_px:
            if current: lines.append(current); current = ""
            for chunk in split_word(word):
                test = (current+" "+chunk).strip()
                if draw_obj.textlength(test, font=font) <= max_px:
                    current = test
                else:
                    if current: lines.append(current)
                    current = chunk
        else:
            test = (current+" "+word).strip()
            if draw_obj.textlength(test, font=font) <= max_px:
                current = test
            else:
                if current: lines.append(current)
                current = word
    if current: lines.append(current)
    return lines or [""]

def draw_label(item_dict, W, H, field_config, global_sup_bytes=None, is_pdf=True):
    """Dibuja un rótulo completo y devuelve una imagen PIL."""
    W, H = max(1, int(W)), max(1, int(H))
    img  = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    # ── Escala de fuentes ─────────────────────────────────────────────────────
    # Siempre usar 595 (ancho A4 en puntos) como referencia
    # para que las proporciones sean iguales en preview y PDF
    ref_w    = 595.0
    scale    = W / ref_w
    title_sz = max(6, int(9  * scale))
    body_sz  = max(6, int(13 * scale))

    title_font = _get_font(title_sz, bold=True)
    bold_font  = _get_font(body_sz,  bold=True)
    body_font  = _get_font(body_sz,  bold=False)

    # ── Encabezado ────────────────────────────────────────────────────────────
    hdr_h = int(H * 0.12)
    draw.rectangle([(0, 0), (W, hdr_h)], fill="#666666")

    text_col_w = W * 0.65
    ref_col_w  = W * 0.23
    sup_col_w  = W * 0.12

    def _draw_hdr_text(text, x_start, col_w):
        tw = draw.textlength(text, font=title_font)
        tx = x_start + (col_w - tw) / 2
        th = title_font.getbbox("A")[3] - title_font.getbbox("A")[1]
        ty = (hdr_h - th) / 2
        draw.text((tx, ty), text, font=title_font, fill="white")

    _draw_hdr_text("DESCRIPCION / MATERIAL", 0,          text_col_w)
    _draw_hdr_text("IMAGEN REFERENCIAL",     text_col_w, ref_col_w)
    _draw_hdr_text("PROVEEDOR",              text_col_w + ref_col_w, sup_col_w)

    # ── Texto de campos ───────────────────────────────────────────────────────
    padding_x = 8
    fc = field_config  # definir fc primero

    # LINE_GAP adaptativo según campos visibles
    visible_fields = sum(1 for k in ['product','oc','fab_date','exp_date',
                                      'batch','quantity','deliver_to',
                                      'description','etq_x_caja']
                         if fc.get(k,{}).get('visible', True))
    if visible_fields >= 7:
        LINE_GAP = max(3, int(5 * scale))
    elif visible_fields >= 5:
        LINE_GAP = max(4, int(6 * scale))
    else:
        LINE_GAP = max(4, int(8 * scale))
    cur_y = hdr_h + 4
    tsw   = int(W * 0.65) - padding_x

    def field_on(key):
        return fc.get(key, {}).get('visible', True)

    def lbl(key, default):
        t = fc.get(key, {}).get('label', default).strip()
        return (t + ': ') if t else ''

    def draw_field(prefix, value, gap=None):
        nonlocal cur_y
        gap = gap or LINE_GAP
        if not prefix and not value: return
        pw = draw.textlength(prefix, font=bold_font) if prefix else 0
        draw.text((padding_x, cur_y), prefix, font=bold_font, fill="black")
        draw.text((padding_x + pw, cur_y), value, font=body_font, fill="black")
        cur_y += max(_line_h(draw, prefix, bold_font),
                     _line_h(draw, value or "A", body_font)) + gap

    # PRODUCTO con fuente adaptativa
    if field_on('product'):
        pfx = lbl('product', 'PRODUCTO')
        txt = item_dict.get('product', '')
        MAX_LINES = 3
        adapted_body = body_font
        adapted_bold = bold_font
        adapted_sz   = body_sz
        while adapted_sz >= max(6, body_sz - 6):
            pw     = draw.textlength(pfx, font=adapted_bold)
            avail  = tsw - pw
            wrapped = _wrap_px(draw, txt, adapted_body, avail)
            if len(wrapped) <= MAX_LINES: break
            adapted_sz  -= 1
            adapted_body = _get_font(adapted_sz, bold=False)
            adapted_bold = _get_font(adapted_sz, bold=True)

        pw      = draw.textlength(pfx, font=adapted_bold)
        avail   = tsw - pw
        wrapped = _wrap_px(draw, txt, adapted_body, avail)
        if wrapped:
            draw.text((padding_x, cur_y), pfx, font=adapted_bold, fill="black")
            draw.text((padding_x + pw, cur_y), wrapped[0], font=adapted_body, fill="black")
            cur_y += _line_h(draw, wrapped[0] or "A", adapted_body) + LINE_GAP
            for ln in wrapped[1:]:
                draw.text((padding_x + pw, cur_y), ln, font=adapted_body, fill="black")
                cur_y += _line_h(draw, ln, adapted_body) + LINE_GAP // 2
        else:
            cur_y += _line_h(draw, "A", body_font) + LINE_GAP

    # OC
    if field_on('oc'):
        draw_field(lbl('oc', 'OC'), item_dict.get('oc', ''))

    # F.ELAB / F.EXP en la misma línea
    show_fab = field_on('fab_date')
    show_exp = field_on('exp_date')
    if show_fab or show_exp:
        lbl_fab = (fc.get('fab_date',{}).get('label','F.ELAB.').strip() + ': ') if show_fab else ''
        lbl_exp = (fc.get('exp_date',{}).get('label','F.EXP.').strip()  + ': ') if show_exp else ''
        cx = padding_x
        if show_fab:
            pw = draw.textlength(lbl_fab, font=bold_font)
            draw.text((cx, cur_y), lbl_fab, font=bold_font, fill="black")
            draw.text((cx + pw, cur_y), item_dict.get('fab_date',''), font=body_font, fill="black")
            cx += pw + draw.textlength(item_dict.get('fab_date',''), font=body_font) + 12
        if show_exp:
            pw = draw.textlength(lbl_exp, font=bold_font)
            draw.text((cx, cur_y), lbl_exp, font=bold_font, fill="black")
            draw.text((cx + pw, cur_y), item_dict.get('exp_date',''), font=body_font, fill="black")
        cur_y += _line_h(draw, "A", body_font) + LINE_GAP

    if field_on('batch'):
        draw_field(lbl('batch', 'Lote'), item_dict.get('batch', ''))
    if field_on('quantity'):
        draw_field(lbl('quantity', 'Cantidad'), item_dict.get('quantity', ''))
    if field_on('deliver_to'):
        draw_field(lbl('deliver_to', 'Entregar'), item_dict.get('deliver_to', ''))
    if field_on('description') and item_dict.get('description','').strip():
        draw_field('', item_dict.get('description',''))
    if field_on('etq_x_caja') and item_dict.get('etq_x_caja','').strip():
        draw_field(lbl('etq_x_caja','ETQ/Caja'), item_dict.get('etq_x_caja',''))

    # ── Imágenes ──────────────────────────────────────────────────────────────
    pad       = 5
    iy1       = hdr_h + pad
    iy2       = H     - pad
    avail_h   = iy2 - iy1

    ref_x1  = int(W * 0.65) + pad
    ref_w   = int(W * 0.23) - 2 * pad
    sup_x1  = int(W * 0.88) + pad
    sup_w   = int(W * 0.12) - 2 * pad

    def paste_img(pil_img, col_x, col_w, max_h_factor=1.0):
        if pil_img is None: return
        thumb = pil_img.copy().convert("RGBA")
        thumb.thumbnail((max(1,col_w), max(1, int(avail_h*max_h_factor))), Image.LANCZOS)
        px = col_x + (col_w - thumb.width)  // 2
        py = iy1   + (avail_h - thumb.height) // 2
        img.paste(thumb, (max(col_x,px), max(iy1,py)), thumb)

    # Imagen referencial
    ref_path = item_dict.get('ref_image_path')
    ref_img  = None
    if ref_path and os.path.exists(ref_path):
        try: ref_img = Image.open(ref_path)
        except Exception: pass
    paste_img(ref_img, ref_x1, ref_w)

    # Imagen proveedor (propia o global)
    sup_path = item_dict.get('sup_image_path')
    sup_img  = None
    if sup_path and os.path.exists(sup_path):
        try: sup_img = Image.open(sup_path)
        except Exception: pass
    if sup_img is None and global_sup_bytes:
        try: sup_img = Image.open(BytesIO(global_sup_bytes))
        except Exception: pass
    paste_img(sup_img, sup_x1, sup_w, max_h_factor=0.70)

    return img

# ══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE PDF
# ══════════════════════════════════════════════════════════════════════════════
def generate_pdf_bytes(selected_items, field_config, global_sup_bytes):
    if not REPORTLAB_OK:
        return None, "ReportLab no instalado. Ejecuta: pip install reportlab"
    try:
        PAGE_W, PAGE_H = portrait(A4)
        MARGIN         = 3 * mm
        USABLE_W       = PAGE_W - 2 * MARGIN
        USABLE_H       = PAGE_H - 2 * MARGIN
        LABEL_H        = 45 * mm
        LABEL_W        = USABLE_W

        # Píxeles para renderizar cada rótulo PIL
        DPI     = 150
        PX_W    = int(LABEL_W / mm * DPI / 25.4)
        PX_H    = int(LABEL_H / mm * DPI / 25.4)

        buf = BytesIO()
        c   = rl_canvas.Canvas(buf, pagesize=portrait(A4))

        # Agrupar en páginas
        pages = [selected_items[i:i+6] for i in range(0, len(selected_items), 6)]

        # Gap fijo de 3mm entre rótulos — no distribuir simétricamente
        # para que no queden flotando con mucho espacio
        GAP_FIXED = 3 * mm

        for page_items in pages:
            n = len(page_items)
            for row, item in enumerate(page_items):
                y_from_top = MARGIN + row * (LABEL_H + GAP_FIXED)
                y_rl       = PAGE_H - y_from_top - LABEL_H

                lbl_img = draw_label(item, PX_W, PX_H, field_config,
                                     global_sup_bytes, is_pdf=True)
                buf_img = BytesIO()
                lbl_img.save(buf_img, format='PNG')
                buf_img.seek(0)
                c.drawImage(ImageReader(buf_img), MARGIN, y_rl,
                            width=LABEL_W, height=LABEL_H)

            c.showPage()

        c.save()
        buf.seek(0)
        return buf.read(), None
    except Exception as e:
        import traceback
        return None, traceback.format_exc()

# ══════════════════════════════════════════════════════════════════════════════
# GUARDAR DATOS
# ══════════════════════════════════════════════════════════════════════════════
def save_items():
    import shutil, datetime as _dt
    # Backup automático (máx 5 copias)
    backup_dir = os.path.join(SCRIPT_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    if os.path.exists(DATA_FILE):
        ts  = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(DATA_FILE, os.path.join(backup_dir, f"label_data_{ts}.json"))
        # Limpiar backups viejos (conservar los últimos 5)
        bkps = sorted(os.listdir(backup_dir))
        for old_bkp in bkps[:-5]:
            try: os.remove(os.path.join(backup_dir, old_bkp))
            except: pass
    _save_json(DATA_FILE, {'items': _get_items()})

def save_image_from_upload(uploaded_file, prefix="img"):
    if uploaded_file is None:
        return None
    ext  = os.path.splitext(uploaded_file.name)[1] or '.png'
    name = f"{prefix}_{uuid.uuid4().hex[:8]}{ext}"
    path = os.path.join(IMAGE_DIR, name)
    with open(path, 'wb') as f:
        f.write(uploaded_file.read())
    return path

def calc_exp_date(fab_str, months):
    try:
        fab = datetime.datetime.strptime(fab_str.strip(), "%d/%m/%Y")
        m   = fab.month + months
        y   = fab.year  + (m - 1) // 12
        m   = (m - 1) % 12 + 1
        d   = min(fab.day, calendar.monthrange(y, m)[1])
        return datetime.date(y, m, d).strftime("%d/%m/%Y")
    except Exception:
        return ""

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR — FORMULARIO
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🏷️ Rotul")

    # ── Selección de rótulo ───────────────────────────────────────────────────
    items = _get_items()
    idx   = st.session_state.selected_idx
    is_new   = (idx is None)
    item_d   = items[idx].copy() if (idx is not None and idx < len(items)) else {}

    # ── Material (calcula F.EXP) ──────────────────────────────────────────────
    mats     = st.session_state.materials
    mat_opts = ["— Sin seleccionar —"] + sorted(mats.keys())
    sel_mat  = st.selectbox("📦 Material (calcula F.EXP)", mat_opts, key="sel_mat")

    st.markdown("---")


    # ── Pins (fijar valores entre rótulos) ───────────────────────────────────
    with st.expander("📌 Fijar campos entre rótulos", expanded=False):
        st.caption("Los campos fijados no se sobreescriben al cargar otro rótulo.")
        pc1, pc2, pc3, pc4 = st.columns(4)
        pin_product  = pc1.checkbox("Producto",   key="pin_product")
        pin_oc       = pc1.checkbox("OC",         key="pin_oc")
        pin_fab      = pc2.checkbox("F.ELAB.",    key="pin_fab")
        pin_exp      = pc2.checkbox("F.EXP.",     key="pin_exp")
        pin_batch    = pc3.checkbox("Lote",       key="pin_batch")
        pin_quantity = pc3.checkbox("Cantidad",   key="pin_qty")
        pin_deliver  = pc4.checkbox("Entregar a", key="pin_deliver")
        pin_copies   = pc4.checkbox("Copias",     key="pin_copies")

    # Valor del campo: si pin activo mantiene lo que está en session_state,
    # si no usa el valor del item seleccionado
    def _pinval(pin_key, pin_active, item_key, default=''):
        if pin_active:
            return st.session_state.get(f"pv_{pin_key}", item_d.get(item_key, default))
        v = item_d.get(item_key, default)
        st.session_state[f"pv_{pin_key}"] = v
        return v

    st.markdown("---")

    # ── Campos del formulario — sin keys en widgets para evitar conflictos ────
    product    = st.text_input("Producto",   value=_pinval("product", pin_product, 'product'))
    oc         = st.text_input("OC",         value=_pinval("oc",      pin_oc,      'oc'))

    col_fab, col_exp = st.columns(2)
    with col_fab:
        fab_date = st.text_input("F. ELAB. (dd/mm/yyyy)",
                                  value=_pinval("fab", pin_fab, 'fab_date'))
    with col_exp:
        exp_auto = ""
        if sel_mat != "— Sin seleccionar —" and fab_date and fab_date.strip():
            exp_auto = calc_exp_date(fab_date, mats.get(sel_mat, 0))
        exp_default = exp_auto if exp_auto else _pinval("exp", pin_exp, 'exp_date')
        exp_date = st.text_input("F. EXP. (dd/mm/yyyy)", value=exp_default)
        if exp_auto:
            st.caption(f"✅ Calculada desde {sel_mat}")

    batch      = st.text_input("Lote",       value=_pinval("batch",   pin_batch,   'batch'))
    quantity   = st.text_input("Cantidad",   value=_pinval("qty",     pin_quantity, 'quantity'))
    deliver_to = st.text_input("Entregar a", value=_pinval("deliver", pin_deliver, 'deliver_to'))
    copies_val = _pinval("copies", pin_copies, 'copies', 1)
    try:    copies_val = max(1, int(copies_val))
    except: copies_val = 1
    copies     = st.number_input("Copias", value=copies_val, min_value=1, max_value=99)
    etq_caja   = st.text_input("ETQ x Caja",  value=item_d.get('etq_x_caja',''))
    description= st.text_area("Descripción",  value=item_d.get('description',''), height=60)

    # ── Imágenes ──────────────────────────────────────────────────────────────
    st.markdown("**🖼 Imagen Referencial**")
    img_tabs = st.tabs(["📂 Subir archivo", "📋 Pegar (Ctrl+V)"])

    ref_upload = None
    ref_paste_img = None

    with img_tabs[0]:
        ref_upload = st.file_uploader("Seleccionar imagen referencial",
                                       type=['png','jpg','jpeg','bmp','gif'],
                                       key="up_ref")

    with img_tabs[1]:
        if PASTE_BTN_OK:
            st.caption("1. Copia una imagen al portapapeles")
            st.caption("2. Haz clic en el botón azul de abajo")
            paste_result = _paste_btn(
                label="📋 Pegar imagen del portapapeles",
                key="paste_ref_btn",
                background_color="#336699",
                hover_background_color="#224488",
                errors="raise"
            )
            if paste_result and paste_result.image_data is not None:
                ref_paste_img = paste_result.image_data
                st.image(ref_paste_img, caption="Vista previa", width=150)
                st.success("✅ Imagen lista — haz clic en Guardar")
        else:
            st.warning("Instala streamlit-paste-button: pip install streamlit-paste-button")

    st.markdown("**🏭 Imagen Proveedor**")
    st.caption("*(Vacío = usa imagen global de proveedor)*")
    sup_upload = st.file_uploader("Seleccionar imagen proveedor",
                                   type=['png','jpg','jpeg','bmp','gif'],
                                   key="up_sup")

    # Miniaturas actuales
    cur_ref = item_d.get('ref_image_path')
    cur_sup = item_d.get('sup_image_path')
    c1, c2  = st.columns(2)
    with c1:
        if cur_ref and os.path.exists(cur_ref):
            st.image(cur_ref, caption="Ref. actual", width=80)
    with c2:
        if cur_sup and os.path.exists(cur_sup):
            st.image(cur_sup, caption="Prov. actual", width=60)

    st.markdown("---")

    # ── Botones de acción ─────────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        if st.button("💾 Guardar", type="primary"):
            new_item = {
                'product':        product,
                'oc':             oc,
                'fab_date':       fab_date,
                'exp_date':       exp_date,
                'batch':          batch,
                'quantity':       quantity,
                'deliver_to':     deliver_to,
                'description':    description,
                'copies':         copies,
                'etq_x_caja':     etq_caja,
                'selected':       item_d.get('selected', False),
                'ref_image_path': item_d.get('ref_image_path'),
                'sup_image_path': item_d.get('sup_image_path'),
            }
            # Imagen referencial: prioridad pegada > subida > existente
            if ref_paste_img is not None:
                # Guardar imagen pegada desde portapapeles
                paste_path = os.path.join(IMAGE_DIR, f"ref_paste_{uuid.uuid4().hex[:8]}.png")
                ref_paste_img.save(paste_path, "PNG")
                new_item['ref_image_path'] = paste_path
            elif ref_upload:
                new_item['ref_image_path'] = save_image_from_upload(ref_upload, "ref")
            # Imagen proveedor
            if sup_upload:
                new_item['sup_image_path'] = save_image_from_upload(sup_upload, "sup")

            cur_items = _get_items()
            if is_new:
                cur_items.append(new_item)
                _set_items(cur_items)
                st.session_state.selected_idx = len(_get_items()) - 1
            else:
                cur_items[idx] = new_item
                _set_items(cur_items)
            save_items()
            st.success("✅ Guardado")
            st.rerun()

    with col2:
        if st.button("🆕 Nuevo"):
            st.session_state.selected_idx = None
            st.rerun()

    col3, col4 = st.columns(2)
    with col3:
        if not is_new and st.button("🗑 Eliminar"):
            cur_items = _get_items()
            cur_items.pop(idx)
            _set_items(cur_items)
            st.session_state.selected_idx = None
            save_items()
            st.rerun()
    with col4:
        if not is_new and st.button("📋 Duplicar"):
            cur_items = _get_items()
            dup = cur_items[idx].copy()
            cur_items.append(dup)
            _set_items(cur_items)
            st.session_state.selected_idx = len(_get_items()) - 1
            save_items()
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ÁREA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
tab_list, tab_preview, tab_config = st.tabs(["📋 Lista de Rótulos", "👁 Vista Previa PDF", "⚙ Configuración"])

# ────────────────────────────────────────────────────────────────────────────
# TAB 1 — LISTA
# ────────────────────────────────────────────────────────────────────────────
with tab_list:
    items = _get_items()

    # Búsqueda y filtros
    f_col1, f_col2 = st.columns([4, 1])
    with f_col1:
        query = st.text_input("🔍 Buscar", key="search_box").lower()
    with f_col2:
        solo_seleccionados = st.checkbox("☑ Solo seleccionados", key="filter_selected")

    # Contadores
    total       = len(items)
    con_img     = sum(1 for i in items if i.get('ref_image_path') or i.get('sup_image_path'))
    sin_lote    = sum(1 for i in items if not i.get('batch','').strip())
    sin_fecha   = sum(1 for i in items if not i.get('fab_date','').strip() or not i.get('exp_date','').strip())
    seleccionados = sum(1 for i in items if i.get('selected', False))

    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("Total",        total)
    m2.metric("🖼 Con imagen", con_img)
    m3.metric("⚠ Sin lote",   sin_lote)
    m4.metric("📅 Sin fecha",  sin_fecha)
    m5.metric("☑ Selec.",     seleccionados)

    st.markdown("---")

    # Botones de selección masiva
    bc1, bc2, bc3, bc4 = st.columns(4)
    def _bulk_sel(value):
        cur = _get_items()
        for idx2, item2 in enumerate(cur):
            v = (not item2.get('selected', False)) if value is None else value
            item2['selected'] = v
            # Forzar el estado visual del checkbox widget
            for k in list(st.session_state.keys()):
                if k.startswith('chk_') and k.endswith('_' + str(idx2)):
                    st.session_state[k] = v
        _set_items(cur)
        save_items()

    with bc1:
        if st.button("☑ Selec. Todo"):
            _bulk_sel(True); st.rerun()
    with bc2:
        if st.button("☐ Deselec. Todo"):
            _bulk_sel(False); st.rerun()
    with bc3:
        if st.button("↔ Invertir"):
            _bulk_sel(None); st.rerun()
    with bc4:
        selected_items = [i for i in items if i.get('selected', False)]
        n_sel = sum(int(i.get('copies',1)) for i in selected_items)
        st.info(f"**{n_sel}** rótulos al PDF")

    st.markdown("---")

    # Ordenar lista
    sort_col = st.selectbox("Ordenar por",
                            ["Producto","Lote","F.Elab.","F.Exp.","Entregar a"],
                            key="sort_col", label_visibility="collapsed",
                            index=0)
    sort_map = {"Producto":"product","Lote":"batch",
                "F.Elab.":"fab_date","F.Exp.":"exp_date","Entregar a":"deliver_to"}
    sort_key = sort_map.get(sort_col, "product")

    # Guardar índice original ANTES de ordenar
    # Seleccionados siempre primero, luego por la columna elegida
    items_with_idx = sorted(enumerate(items),
                            key=lambda x: (
                                0 if x[1].get('selected', False) else 1,
                                x[1].get(sort_key,'').lower()
                            ))

    # Filtrar por búsqueda y por seleccionados
    filtered = [
        (orig_idx, it) for orig_idx, it in items_with_idx
        if (not query
            or query in it.get('product','').lower()
            or query in it.get('batch','').lower()
            or query in it.get('oc','').lower()
            or query in it.get('deliver_to','').lower())
        and (not solo_seleccionados or it.get('selected', False))
    ]

    if not filtered:
        st.info("No hay rótulos. Usa el formulario de la izquierda para agregar uno.")
    else:
        # Encabezado de la lista
        h1, h2, h3 = st.columns([0.5, 8, 1.5])
        h1.caption("✓")
        h2.caption(f"**{len(filtered)}** rótulos encontrados")

    for display_pos, (real_idx, it) in enumerate(filtered):
        col_chk, col_info, col_btn = st.columns([0.5, 8, 1.5])
        col_chk, col_thumb, col_info = st.columns([0.4, 1.0, 8.6])
        with col_chk:
            checked = st.checkbox("", value=it.get('selected', False),
                                  key=f"chk_{display_pos}_{real_idx}",
                                  label_visibility="collapsed")
            if checked != it.get('selected', False):
                cur = _get_items()
                if real_idx < len(cur):
                    cur[real_idx]['selected'] = checked
                    _set_items(cur)
                    save_items()
                    st.rerun()
        with col_thumb:
            ref_path = it.get('ref_image_path')
            if ref_path and os.path.exists(ref_path):
                try:
                    # Miniatura del mismo alto que ~2 líneas de texto (~44px)
                    st.image(ref_path, width=60)
                except Exception:
                    st.markdown('<div style="width:60px;height:44px;background:#eee;border-radius:4px;display:flex;align-items:center;justify-content:center;color:#bbb;font-size:18px;">📷</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div style="width:60px;height:44px;background:#f5f5f5;border:1px dashed #ccc;border-radius:4px;display:flex;align-items:center;justify-content:center;color:#ccc;font-size:18px;">📷</div>', unsafe_allow_html=True)
        with col_info:
            # Info y botones en la misma columna
            impreso  = " ✅" if it.get('printed', False) else ""
            lote_txt = f"Lote: {it.get('batch','')}  " if it.get('batch','').strip() else ""
            exp_txt  = f"F.EXP: {it.get('exp_date','')}  " if it.get('exp_date','').strip() else ""
            prod_txt = it.get("product", "(sin producto)")
            info_txt = f"{lote_txt}{exp_txt}x{it.get('copies',1)}{impreso}"
            # Texto e info en la mitad izquierda, botones en la derecha
            ci1, ci2 = st.columns([7, 1])
            with ci1:
                st.markdown(f"**{prod_txt}**  \n{info_txt}", unsafe_allow_html=True)
            with ci2:
                if st.button("✏️", key=f"edit_{display_pos}_{real_idx}", help="Editar"):
                    st.session_state.selected_idx = real_idx
                    st.rerun()
                is_printed = it.get('printed', False)
                if st.button("✅" if is_printed else "🖨️",
                             key=f"print_{display_pos}_{real_idx}",
                             help="Marcar como NO impreso" if is_printed else "Marcar como impreso"):
                    cur = _get_items()
                    cur[real_idx]['printed'] = not is_printed
                    _set_items(cur)
                    save_items()
                    st.rerun()
        st.divider()
# ────────────────────────────────────────────────────────────────────────────
# TAB 2 — VISTA PREVIA + GENERAR PDF
# ────────────────────────────────────────────────────────────────────────────
with tab_preview:
    _items_safe = [i for i in st.session_state.get('_label_items', [])
                   if isinstance(i, dict)]
    selected_items_all = []
    for it in _items_safe:
        if it.get('selected', False):
            try: copies = max(1, int(it.get('copies', 1)))
            except: copies = 1
            for _ in range(copies):
                selected_items_all.append(it)

    max_per_page   = 6
    total_pages    = max(1, (len(selected_items_all) + max_per_page - 1) // max_per_page) if selected_items_all else 1
    preview_page   = st.session_state.get('preview_page', 0)
    preview_page   = max(0, min(preview_page, total_pages - 1))

    # Navegación
    pn1, pn2, pn3, pn4, pn5 = st.columns([1,1,3,1,1])
    with pn1:
        if st.button("⏮"):
            st.session_state.preview_page = 0; st.rerun()
    with pn2:
        if st.button("◀"):
            st.session_state.preview_page = max(0, preview_page-1); st.rerun()
    with pn3:
        st.markdown(f"<center>Página **{preview_page+1}** de **{total_pages}**</center>",
                    unsafe_allow_html=True)
    with pn4:
        if st.button("▶"):
            st.session_state.preview_page = min(total_pages-1, preview_page+1); st.rerun()
    with pn5:
        if st.button("⏭"):
            st.session_state.preview_page = total_pages-1; st.rerun()

    # Renderizar página actual
    page_items = selected_items_all[preview_page*max_per_page:(preview_page+1)*max_per_page]

    zoom = st.slider("🔍 Zoom", min_value=50, max_value=150, value=100, step=10,
                     format="%d%%", key="preview_zoom")

    if not page_items:
        st.info("Selecciona rótulos en la lista para ver la vista previa.")
    else:
        # Crear imagen de hoja A4
        # Usamos resolución alta para que el texto quepa bien
        zoom_factor = zoom / 100
        PREVIEW_DPI = 150
        MM_TO_PX    = PREVIEW_DPI / 25.4

        # Dimensiones reales del A4 y del rótulo en píxeles a ese DPI
        A4_W_PX   = int(210 * MM_TO_PX * zoom_factor)  # 210mm
        MARGIN_PX = int(3   * MM_TO_PX * zoom_factor)  # 3mm margen
        lbl_w     = A4_W_PX - 2 * MARGIN_PX
        lbl_h     = int(45  * MM_TO_PX * zoom_factor)  # 45mm por rótulo
        GAP_PX    = int(3   * MM_TO_PX * zoom_factor)  # 3mm gap fijo

        # Alto total de la hoja
        n = len(page_items)
        sheet_h = MARGIN_PX + n * lbl_h + (n - 1) * GAP_PX + MARGIN_PX

        sheet = Image.new("RGB", (A4_W_PX, sheet_h), "white")
        sheet_draw = ImageDraw.Draw(sheet)
        sheet_draw.rectangle([(0, 0), (A4_W_PX - 1, sheet_h - 1)],
                              outline="#3366cc", width=2)

        fc  = st.session_state.field_config
        gsb = st.session_state.global_sup_bytes

        for i, it in enumerate(page_items):
            y0 = MARGIN_PX + i * (lbl_h + GAP_PX)
            lbl_img = draw_label(it, lbl_w, lbl_h, fc, gsb, is_pdf=False)
            sheet.paste(lbl_img, (MARGIN_PX, y0))
            sheet_draw.rectangle([(MARGIN_PX, y0), (MARGIN_PX+lbl_w, y0+lbl_h)],
                                  outline="black", width=1)

        st.image(sheet, use_container_width=True)

    st.markdown("---")

    # Generar PDF
    if selected_items_all:
        n_rotulos = len(selected_items_all)
        n_paginas = (n_rotulos + 5) // 6
        st.info(f"📋 **{n_rotulos}** rótulos → **{n_paginas}** página{'s' if n_paginas>1 else ''} de PDF")

    if st.button("📄 Generar y Descargar PDF", type="primary", disabled=not selected_items_all):
        prog = st.progress(0, text="Preparando PDF...")
        with st.spinner("Generando PDF..."):
            prog.progress(30, text="Renderizando rótulos...")
            pdf_bytes, err = generate_pdf_bytes(
                selected_items_all,
                st.session_state.field_config,
                st.session_state.global_sup_bytes
            )
            prog.progress(100, text="¡Listo!")
        if err:
            st.error(f"Error: {err}")
        else:
            fname = f"rotulos_{datetime.date.today().strftime('%Y%m%d')}.pdf"
            st.download_button("⬇️ Descargar PDF", data=pdf_bytes,
                               file_name=fname, mime="application/pdf")

# ────────────────────────────────────────────────────────────────────────────
# TAB 3 — CONFIGURACIÓN GLOBAL
# ────────────────────────────────────────────────────────────────────────────
with tab_config:
    st.markdown("### ⚙ Configuración Global")

    cfg_col1, cfg_col2 = st.columns(2)

    # ── Imagen global de proveedor ────────────────────────────────────────────
    with cfg_col1:
        st.markdown("#### 🏭 Imagen de Proveedor Global")
        if st.session_state.global_sup_bytes:
            st.image(BytesIO(st.session_state.global_sup_bytes), width=120)
            if st.button("🗑 Quitar imagen global"):
                st.session_state.global_sup_bytes = None
                if os.path.exists(GLOBAL_SUPPLIER_PATH):
                    os.remove(GLOBAL_SUPPLIER_PATH)
                st.success("Imagen global eliminada")
                st.rerun()
        else:
            st.info("Sin imagen global")
        new_global = st.file_uploader("📂 Cargar imagen global",
                                       type=['png','jpg','jpeg','bmp'],
                                       key="up_global")
        if new_global:
            img_bytes = new_global.read()
            with open(GLOBAL_SUPPLIER_PATH, 'wb') as f:
                f.write(img_bytes)
            st.session_state.global_sup_bytes = img_bytes
            st.success("✅ Imagen global guardada")
            st.rerun()

    # ── Campos visibles ───────────────────────────────────────────────────────
    with cfg_col2:
        st.markdown("#### 📋 Campos visibles en el rótulo")
        fc    = st.session_state.field_config
        changed = False
        for key in ['product','oc','fab_date','exp_date','batch',
                    'quantity','deliver_to','description','etq_x_caja']:
            cc1, cc2 = st.columns([1, 3])
            with cc1:
                vis = st.checkbox("", value=fc[key]['visible'], key=f"fc_vis_{key}")
                if vis != fc[key]['visible']:
                    fc[key]['visible'] = vis; changed = True
            with cc2:
                lbl_txt = st.text_input("", value=fc[key]['label'],
                                         key=f"fc_lbl_{key}",
                                         label_visibility="collapsed")
                if lbl_txt != fc[key]['label']:
                    fc[key]['label'] = lbl_txt; changed = True
        if changed:
            _save_json(FIELD_CONFIG_PATH, fc)

    st.markdown("---")

    # ── Materiales y vida útil ────────────────────────────────────────────────
    st.markdown("#### 📦 Materiales y Vida Útil")
    mats = st.session_state.materials

    mat_c1, mat_c2, mat_c3 = st.columns([3, 1, 1])
    with mat_c1:
        new_mat_name = st.text_input("Nombre del material", key="new_mat_name")
    with mat_c2:
        new_mat_months = st.number_input("Meses", min_value=1, max_value=360,
                                          value=12, key="new_mat_months")
    with mat_c3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ Agregar material"):
            if new_mat_name.strip():
                mats[new_mat_name.strip()] = int(new_mat_months)
                _save_json(MATERIALS_CONFIG_PATH, mats)
                st.success(f"'{new_mat_name}' agregado")
                st.rerun()

    if mats:
        st.markdown("**Materiales guardados:**")
        for name, months in sorted(mats.items()):
            mc1, mc2, mc3 = st.columns([4, 1, 1])
            mc1.write(f"**{name}**")
            mc2.write(f"{months} meses")
            with mc3:
                if st.button("🗑", key=f"del_mat_{name}"):
                    del mats[name]
                    _save_json(MATERIALS_CONFIG_PATH, mats)
                    st.rerun()
    else:
        st.info("No hay materiales configurados.")

    st.markdown("---")

    # ── Importar Pedido (Excel de pedidos) ────────────────────────────────────
    st.markdown("#### 📋 Importar desde Pedido")
    st.caption("Sube el archivo de pedidos para generar rótulos automáticamente.")

    pedido_file = st.file_uploader("Seleccionar archivo de pedido",
                                    type=['xlsx','xls'], key="import_pedido")

    if pedido_file:
        try:
            import openpyxl as _xl
            import re as _re
            import datetime as _dt

            wb  = _xl.load_workbook(pedido_file)
            ws  = wb.active
            hdrs = [str(c.value or '').strip() for c in ws[1]]

            # Leer filas
            pedido_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                rd = {hdrs[i]: row[i] for i in range(len(hdrs)) if i < len(row)}
                pedido_rows.append(rd)

            st.info(f"📄 {len(pedido_rows)} líneas encontradas en el pedido")

            # ── Formulario de parámetros ──────────────────────────────────────
            with st.form("form_pedido"):
                st.markdown("**Parámetros del pedido:**")
                fab_date_ped = st.text_input("📅 Fecha de emisión (dd/mm/yyyy)",
                                              placeholder="ej: 07/05/2026")

                st.markdown("---")
                st.markdown("**Vista previa de líneas a procesar:**")

                # Previsualizar qué se va a crear
                base_items = _get_items()
                preview_data = []

                for rd in pedido_rows:
                    pedido_num = str(rd.get('Pedido','') or '').strip()
                    producto   = str(rd.get('Producto','') or '').strip()
                    material   = str(rd.get('Material','') or '').strip()
                    cantidad   = str(rd.get('Cantidad','') or '').strip()
                    cliente    = str(rd.get('Cliente','') or '').strip()

                    # Extraer número de pedido (sin P)
                    num_match = _re.match(r'P?(\d+)', pedido_num)
                    num_ped   = num_match.group(1) if num_match else pedido_num

                    # Extraer código del producto (ej: NES-298-235)
                    cod_match = _re.match(r'^([A-Z0-9]+-\d+-\d+)', producto, _re.IGNORECASE)
                    cod_prod  = cod_match.group(1).upper() if cod_match else ''

                    # Buscar en base de rótulos
                    found_item = None
                    for it in base_items:
                        prod_code = _re.match(r'^([A-Z0-9]+-\d+-\d+)', it.get('product',''), _re.IGNORECASE)
                        if prod_code and prod_code.group(1).upper() == cod_prod:
                            found_item = it
                            break

                    # Buscar material en config — búsqueda flexible por palabras
                    mats = st.session_state.get('materials', {})
                    mat_months  = None
                    mat_matched = None
                    mat_upper   = material.upper().strip()
                    # 1. Coincidencia exacta
                    for mat_name, months in mats.items():
                        if mat_name.upper().strip() == mat_upper:
                            mat_months = months; mat_matched = mat_name; break
                    # 2. Contiene
                    if mat_months is None:
                        for mat_name, months in mats.items():
                            if mat_name.upper() in mat_upper or mat_upper in mat_name.upper():
                                mat_months = months; mat_matched = mat_name; break
                    # 3. Palabras en común (al menos 1 palabra clave)
                    if mat_months is None:
                        mat_words = set(mat_upper.split())
                        best_score, best_match = 0, None
                        for mat_name, months in mats.items():
                            cfg_words = set(mat_name.upper().split())
                            score = len(mat_words & cfg_words)
                            if score > best_score:
                                best_score = score; best_match = (mat_name, months)
                        if best_score >= 1 and best_match:
                            mat_months = best_match[1]; mat_matched = best_match[0]

                    preview_data.append({
                        'pedido_num': pedido_num,
                        'num_ped':    num_ped,
                        'producto':   producto,
                        'cod_prod':   cod_prod,
                        'material':   material,
                        'cantidad':   cantidad,
                        'cliente':    cliente,
                        'found_item': found_item,
                        'mat_months': mat_months,
                    })

                # Mostrar tabla de preview
                for pd_row in preview_data:
                    col_st, col_info = st.columns([1, 6])
                    if pd_row['found_item']:
                        col_st.success("✅")
                        col_info.write(f"**{pd_row['pedido_num']}** — {pd_row['cod_prod']} "
                                      f"— Material: {pd_row['material']} "
                                      f"({'%d meses' % pd_row['mat_months'] if pd_row['mat_months'] else '⚠ material no en config'})")
                    else:
                        col_st.error("❌")
                        col_info.write(f"**{pd_row['pedido_num']}** — `{pd_row['cod_prod']}` "
                                      f"**NO ENCONTRADO en base** — {pd_row['producto'][:60]}")

                submitted = st.form_submit_button("✅ Generar Rótulos", type="primary")

            if submitted:
                if not fab_date_ped.strip():
                    st.error("⚠ Ingresa la fecha de emisión")
                else:
                    # Parsear fecha
                    try:
                        fab_dt = _dt.datetime.strptime(fab_date_ped.strip(), "%d/%m/%Y")
                        fab_str = fab_dt.strftime("%d/%m/%Y")
                        fecha_corta = fab_dt.strftime("%d%m%Y")
                    except ValueError:
                        st.error("⚠ Formato de fecha incorrecto. Usa dd/mm/yyyy")
                        st.stop()

                    generados   = []
                    no_encontrados = []

                    for pd_row in preview_data:
                        if not pd_row['found_item']:
                            # Crear rótulo nuevo con datos del pedido
                            base = {
                                'product':        pd_row['producto'],
                                'oc':             '',
                                'fab_date':       '',
                                'exp_date':       '',
                                'batch':          '',
                                'quantity':       '',
                                'deliver_to':     '',
                                'description':    '⚠ NUEVO — agregar imagen',
                                'copies':         1,
                                'etq_x_caja':     '',
                                'selected':       True,
                                'ref_image_path': None,
                                'sup_image_path': None,
                            }
                            no_encontrados.append(pd_row)
                        else:
                            base = pd_row['found_item'].copy()

                        # Lote = número pedido + fecha emisión
                        lote = pd_row['num_ped'] + fecha_corta

                        # F.EXP calculada desde material
                        exp_str = ''
                        if pd_row['mat_months'] and fab_str:
                            exp_str = calc_exp_date(fab_str, pd_row['mat_months'])

                        new_rotulo = {
                            'product':        base.get('product', pd_row['producto']),
                            'oc':             base.get('oc', ''),
                            'fab_date':       fab_str,
                            'exp_date':       exp_str,
                            'batch':          lote,
                            'quantity':       '',
                            'deliver_to':     base.get('deliver_to', ''),
                            'description':    base.get('description', ''),
                            'copies':         1,
                            'etq_x_caja':     base.get('etq_x_caja', ''),
                            'selected':       True,
                            'ref_image_path': base.get('ref_image_path'),
                            'sup_image_path': base.get('sup_image_path'),
                        }
                        generados.append(new_rotulo)

                        # Si es nuevo, también agregarlo al catálogo base
                        if not pd_row['found_item']:
                            catalogo_nuevo = new_rotulo.copy()
                            catalogo_nuevo['selected'] = False
                            catalogo_nuevo['batch']    = ''
                            catalogo_nuevo['fab_date'] = ''
                            catalogo_nuevo['exp_date'] = ''
                            catalogo_nuevo['description'] = '⚠ NUEVO — agregar imagen'
                            # Solo agregarlo si no existe ya
                            cur_base = _get_items()
                            ya_existe = any(
                                _re.match(r'^([A-Z0-9]+-\d+-\d+)', it.get('product',''), _re.IGNORECASE) and
                                _re.match(r'^([A-Z0-9]+-\d+-\d+)', it.get('product','')).group(1).upper() == pd_row['cod_prod']
                                for it in cur_base
                                if _re.match(r'^([A-Z0-9]+-\d+-\d+)', it.get('product',''), _re.IGNORECASE)
                            )
                            if not ya_existe:
                                cur_base.append(catalogo_nuevo)
                                _set_items(cur_base)

                    if generados:
                        cur = _get_items()
                        cur.extend(generados)
                        _set_items(cur)
                        save_items()
                        # Guardar en historial
                        hist = _load_json(HISTORY_PATH, [])
                        hist.insert(0, {
                            'fecha':     datetime.date.today().strftime("%d/%m/%Y"),
                            'hora':      datetime.datetime.now().strftime("%H:%M"),
                            'generados': len(generados),
                            'nuevos':    len(no_encontrados),
                            'fab_date':  fab_date_ped,
                        })
                        _save_json(HISTORY_PATH, hist[:50])  # máx 50 registros
                        st.success(f"✅ {len(generados)} rótulos generados y seleccionados")

                    if no_encontrados:
                        st.warning(f"⚠ {len(no_encontrados)} productos nuevos creados (sin imagen):")
                        for nr in no_encontrados:
                            st.write(f"• **{nr['pedido_num']}** — `{nr['cod_prod']}` — {nr['producto'][:70]}")
                        st.info("Búscalos en la lista con el filtro ⚠ y agrégales la imagen referencial.")

        except Exception as e:
            st.error(f"Error al procesar pedido: {e}")

    st.markdown("---")

    # ── Historial de importaciones ────────────────────────────────────────────
    st.markdown("#### 📜 Historial de Pedidos Importados")
    hist_data = _load_json(HISTORY_PATH, [])
    if hist_data:
        for h in hist_data[:10]:
            st.markdown(
                f"📅 **{h.get('fecha','')} {h.get('hora','')}** — "
                f"F.Elab: `{h.get('fab_date','')}` — "
                f"✅ {h.get('generados',0)} generados / "
                f"🆕 {h.get('nuevos',0)} nuevos"
            )
    else:
        st.caption("Sin historial aún.")

    st.markdown("---")

    # ── Importar desde Excel ──────────────────────────────────────────────────
    st.markdown("#### 📥 Importar desde Excel")
    st.caption("El Excel debe tener columnas: product, oc, fab_date, exp_date, batch, quantity, deliver_to, description, copies, etq_x_caja")
    excel_upload = st.file_uploader("Seleccionar archivo Excel",
                                     type=['xlsx','xls'], key="import_excel")
    col_imp1, col_imp2 = st.columns(2)
    with col_imp1:
        replace_mode = st.checkbox("Reemplazar todo (si no, agrega al final)", value=False)
    with col_imp2:
        if excel_upload and st.button("📥 Importar Excel"):
            try:
                import openpyxl as _xl
                wb  = _xl.load_workbook(excel_upload)
                ws  = wb.active
                hdrs = [str(c.value or '').strip().lower() for c in ws[1]]
                field_map = {
                    'product':'product','oc':'oc','fab_date':'fab_date',
                    'exp_date':'exp_date','batch':'batch','quantity':'quantity',
                    'deliver_to':'deliver_to','description':'description',
                    'copies':'copies','etq_x_caja':'etq_x_caja'
                }
                new_items = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    item = {k: '' for k in field_map}
                    item.update({'copies':1,'selected':False,
                                 'ref_image_path':None,'sup_image_path':None})
                    for ci, hdr in enumerate(hdrs):
                        if hdr in field_map and ci < len(row):
                            val = row[ci]
                            if hdr == 'copies':
                                try: item['copies'] = max(1, int(val))
                                except: item['copies'] = 1
                            else:
                                item[hdr] = str(val or '').strip()
                    if item.get('product'):
                        new_items.append(item)
                cur = [] if replace_mode else _get_items()
                cur.extend(new_items)
                _set_items(cur)
                save_items()
                st.success(f"✅ {len(new_items)} productos importados")
                st.rerun()
            except Exception as e:
                st.error(f"Error al importar: {e}")

    st.markdown("---")

    # ── Exportar datos ────────────────────────────────────────────────────────
    st.markdown("#### 💾 Exportar datos")
    exp_c1, exp_c2 = st.columns(2)
    with exp_c1:
        if [i for i in (st.session_state.get('_label_items',[]) if not callable(st.session_state.get('_label_items')) else []) if isinstance(i,dict)]:
            csv_lines = ["producto,oc,fecha_elab,fecha_exp,lote,cantidad,entregar,copias"]
            for it in [i for i in (st.session_state.get('_label_items',[]) if not callable(st.session_state.get('_label_items')) else []) if isinstance(i,dict)]:
                csv_lines.append(
                    f"{it.get('product','')},{it.get('oc','')},{it.get('fab_date','')},"
                    f"{it.get('exp_date','')},{it.get('batch','')},{it.get('quantity','')},"
                    f"{it.get('deliver_to','')},{it.get('copies',1)}"
                )
            st.download_button("📥 Exportar CSV",
                               data="\n".join(csv_lines).encode('utf-8'),
                               file_name="rotulos_export.csv",
                               mime="text/csv")
    with exp_c2:
        _items_exp = [i for i in st.session_state.get('_label_items', [])
                      if isinstance(i, dict)]
        json_str = json.dumps(_items_exp, indent=2, ensure_ascii=False)
        st.download_button("📥 Exportar JSON",
                           data=json_str.encode('utf-8'),
                           file_name="rotulos_backup.json",
                           mime="application/json")
